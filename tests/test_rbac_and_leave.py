import asyncio
import unittest
from datetime import datetime, time, timedelta
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database.base import Base
from app.models.attendance import AttendanceRecord
from app.models.event import Event
from app.models.leave import LeaveRequest
from app.models.menu import Menu
from app.models.notification import Notification
from app.models.permission import Permission, UserPermission
from app.models.qr import QRToken
from app.models.room import Room, RoomReservation
from app.models.user import User
from app.core.rbac import normalize_role, scoped_user_ids
from app.core.timezone import turkey_now, turkey_today
from app.routers.attendance import export_attendance_excel
from app.routers.event import create_event, send_due_event_reminders
from app.routers.leave import (
    approve_leave_request,
    create_leave_request,
    delete_leave_request,
    get_team_leaves,
)
from app.routers.permission import assign_permission_to_user, remove_permission_from_user
from app.routers.menu import purge_old_menus
from app.routers.user import delete_user, purge_inactive_users
from app.schemas.event import EventCreate
from app.schemas.leave import LeaveCreate
from app.schemas.permission import UserPermissionCreate
from app.core.security import hash_password


class RbacAndLeaveTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        testing_session = sessionmaker(
            autocommit=False,
            autoflush=False,
            bind=self.engine,
        )
        Base.metadata.create_all(bind=self.engine)
        self.db = testing_session()

    def tearDown(self):
        self.db.close()
        Base.metadata.drop_all(bind=self.engine)

    def add_user(self, full_name, email, role="employee", supervisor_id=None):
        user = User(
            full_name=full_name,
            email=email,
            hashed_password=hash_password("secret"),
            role=role,
            supervisor_id=supervisor_id,
        )
        self.db.add(user)
        self.db.commit()
        self.db.refresh(user)
        return user

    def test_normalize_role_accepts_legacy_employee_alias(self):
        self.assertEqual(normalize_role("çalışan"), "employee")
        self.assertEqual(normalize_role("calisan"), "employee")
        self.assertEqual(normalize_role("employee"), "employee")
        self.assertEqual(normalize_role("superadmin"), "super_admin")

    def test_scoped_user_ids_match_role_hierarchy(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        admin = self.add_user("Admin", "admin@example.com", "admin")
        employee = self.add_user(
            "Employee",
            "employee@example.com",
            "employee",
            admin.id,
        )
        other = self.add_user("Other", "other@example.com", "employee")

        self.assertEqual(
            set(scoped_user_ids(self.db, super_admin)),
            {super_admin.id, admin.id, employee.id, other.id},
        )
        self.assertEqual(scoped_user_ids(self.db, admin), [employee.id])
        self.assertEqual(scoped_user_ids(self.db, employee), [employee.id])

    def test_admin_can_approve_subordinate_leave_and_notify_employee(self):
        admin = self.add_user("Admin", "admin@example.com", "admin")
        employee = self.add_user(
            "Employee",
            "employee@example.com",
            "employee",
            admin.id,
        )
        leave = LeaveRequest(
            user_id=employee.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Test",
            status="pending",
        )
        self.db.add(leave)
        self.db.commit()
        self.db.refresh(leave)

        response = approve_leave_request(
            leave.id,
            {"sub": admin.email, "role": admin.role},
            self.db,
        )

        notification = self.db.query(Notification).filter(
            Notification.user_id == employee.id
        ).first()

        self.assertEqual(response["status"], "approved")
        self.assertIsNotNone(notification)
        self.assertEqual(notification.link, "/leaves")

    def test_leave_request_notifies_actual_approvers(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        admin = self.add_user("Admin", "admin@example.com", "admin")
        unrelated_admin = self.add_user(
            "Other Admin",
            "other-admin@example.com",
            "admin",
        )
        permission_user = self.add_user(
            "Permission User",
            "permission@example.com",
            "employee",
        )
        employee = self.add_user(
            "Employee",
            "employee@example.com",
            "employee",
            admin.id,
        )
        permission = Permission(
            code="leave.approve",
            description="Leave approval",
        )
        self.db.add(permission)
        self.db.commit()
        self.db.add(UserPermission(
            user_id=permission_user.id,
            permission_id=permission.id,
        ))
        self.db.commit()

        create_leave_request(
            LeaveCreate(
                start_time=datetime.utcnow(),
                end_time=datetime.utcnow() + timedelta(hours=8),
                reason="Test",
            ),
            {"sub": employee.email, "role": employee.role},
            self.db,
        )

        notified_user_ids = {
            notification.user_id
            for notification in self.db.query(Notification).filter(
                Notification.link == "/leaves"
            ).all()
        }

        self.assertIn(super_admin.id, notified_user_ids)
        self.assertIn(admin.id, notified_user_ids)
        self.assertIn(permission_user.id, notified_user_ids)
        self.assertNotIn(unrelated_admin.id, notified_user_ids)

    def test_direct_supervisor_receives_and_can_approve_leave(self):
        supervisor = self.add_user(
            "Supervisor",
            "supervisor@example.com",
            "employee",
        )
        employee = self.add_user(
            "Employee",
            "employee-direct@example.com",
            "employee",
            supervisor.id,
        )

        create_leave_request(
            LeaveCreate(
                start_time=datetime.utcnow(),
                end_time=datetime.utcnow() + timedelta(hours=8),
                reason="Direct supervisor test",
            ),
            {"sub": employee.email, "role": employee.role},
            self.db,
        )

        notification = self.db.query(Notification).filter(
            Notification.user_id == supervisor.id,
            Notification.link == "/leaves",
        ).first()
        leave = self.db.query(LeaveRequest).filter(
            LeaveRequest.user_id == employee.id,
        ).first()

        self.assertIsNotNone(notification)
        self.assertIsNotNone(leave)

        response = approve_leave_request(
            leave.id,
            {"sub": supervisor.email, "role": supervisor.role},
            self.db,
        )

        self.assertEqual(response["status"], "approved")

    def test_weekly_and_report_leaves_are_auto_approved_without_notifications(self):
        supervisor = self.add_user("Supervisor", "auto-supervisor@example.com", "admin")
        employee = self.add_user(
            "Employee",
            "auto-employee@example.com",
            "employee",
            supervisor.id,
        )

        for leave_type in ["weekly", "report"]:
            response = create_leave_request(
                LeaveCreate(
                    start_time=datetime.utcnow(),
                    end_time=datetime.utcnow() + timedelta(hours=8),
                    leave_type=leave_type,
                ),
                {"sub": employee.email, "role": employee.role},
                self.db,
            )
            leave = self.db.query(LeaveRequest).filter(
                LeaveRequest.id == response["leave_request_id"],
            ).first()

            self.assertEqual(leave.status, "approved")
            self.assertEqual(leave.leave_type, leave_type)

        self.assertEqual(
            self.db.query(Notification).filter(Notification.link == "/leaves").count(),
            0,
        )

    def test_weekly_leave_is_limited_to_two_days_per_month(self):
        employee = self.add_user(
            "Employee",
            "weekly-limit@example.com",
            "employee",
        )

        first_day = datetime(2026, 8, 3, 0, 0)
        second_day = datetime(2026, 8, 10, 0, 0)
        third_day = datetime(2026, 8, 17, 0, 0)

        for leave_day in [first_day, second_day]:
            create_leave_request(
                LeaveCreate(
                    start_time=leave_day,
                    end_time=leave_day.replace(hour=23, minute=59),
                    leave_type="weekly",
                ),
                {"sub": employee.email, "role": employee.role},
                self.db,
            )

        with self.assertRaises(HTTPException) as context:
            create_leave_request(
                LeaveCreate(
                    start_time=third_day,
                    end_time=third_day.replace(hour=23, minute=59),
                    leave_type="weekly",
                ),
                {"sub": employee.email, "role": employee.role},
                self.db,
            )

        self.assertEqual(context.exception.status_code, 400)

    def test_admin_cannot_approve_non_subordinate_leave(self):
        admin = self.add_user("Admin", "admin@example.com", "admin")
        employee = self.add_user("Employee", "employee@example.com", "employee")
        leave = LeaveRequest(
            user_id=employee.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Test",
            status="pending",
        )
        self.db.add(leave)
        self.db.commit()
        self.db.refresh(leave)

        with self.assertRaises(HTTPException) as context:
            approve_leave_request(
                leave.id,
                {"sub": admin.email, "role": admin.role},
                self.db,
            )

        self.assertEqual(context.exception.status_code, 403)

    def test_super_admin_can_approve_own_leave(self):
        super_admin = self.add_user("Super", "super-own@example.com", "super_admin")
        leave = LeaveRequest(
            user_id=super_admin.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Own leave",
            status="pending",
        )
        self.db.add(leave)
        self.db.commit()
        self.db.refresh(leave)

        team_leaves = get_team_leaves(
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        response = approve_leave_request(
            leave.id,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )

        self.db.refresh(leave)

        self.assertIn(leave.id, {item["id"] for item in team_leaves["leaves"]})
        self.assertEqual(response["status"], "approved")
        self.assertEqual(leave.approved_by, super_admin.id)

    def test_user_can_delete_own_approved_leave(self):
        employee = self.add_user("Employee", "employee@example.com", "employee")
        leave = LeaveRequest(
            user_id=employee.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Test",
            status="approved",
        )
        self.db.add(leave)
        self.db.commit()
        self.db.refresh(leave)

        response = delete_leave_request(
            leave.id,
            {"sub": employee.email, "role": employee.role},
            self.db,
        )

        self.assertEqual(response["leave_id"], leave.id)
        self.assertIsNone(
            self.db.query(LeaveRequest).filter(LeaveRequest.id == leave.id).first()
        )

    def test_delete_user_removes_admin_and_related_records(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        admin = self.add_user("Admin", "admin@example.com", "admin")
        employee = self.add_user(
            "Employee",
            "employee@example.com",
            "employee",
            admin.id,
        )
        leave = LeaveRequest(
            user_id=admin.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Test",
            status="pending",
        )
        approved_leave = LeaveRequest(
            user_id=employee.id,
            start_time=datetime.utcnow(),
            end_time=datetime.utcnow() + timedelta(hours=8),
            reason="Approved",
            status="approved",
            approved_by=admin.id,
        )
        attendance = AttendanceRecord(user_id=admin.id, record_type="check_in")
        notification = Notification(
            user_id=admin.id,
            title="Test",
            message="Test",
        )
        permission = Permission(code="attendance.view", description="Attendance")
        event = Event(
            title="Event",
            start_time=datetime.utcnow(),
            created_by=admin.id,
        )
        room = Room(name="Room")

        self.db.add_all([
            leave,
            approved_leave,
            attendance,
            notification,
            permission,
            event,
            room,
        ])
        self.db.commit()
        self.db.refresh(permission)
        self.db.refresh(room)

        user_permission = UserPermission(
            user_id=admin.id,
            permission_id=permission.id,
        )
        room_reservation = RoomReservation(
            room_id=room.id,
            title="Reservation",
            start_date=datetime.utcnow().date(),
            end_date=datetime.utcnow().date(),
            weekday=0,
            start_time=datetime.utcnow().time(),
            end_time=(datetime.utcnow() + timedelta(hours=1)).time(),
            created_by=admin.id,
        )
        self.db.add_all([user_permission, room_reservation])
        self.db.commit()

        response = delete_user(
            admin.id,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )

        self.db.refresh(employee)
        self.db.refresh(approved_leave)

        self.assertEqual(response["user_id"], admin.id)
        self.assertIsNone(self.db.query(User).filter(User.id == admin.id).first())
        self.assertIsNone(employee.supervisor_id)
        self.assertIsNone(approved_leave.approved_by)
        self.assertEqual(
            self.db.query(AttendanceRecord).filter(
                AttendanceRecord.user_id == admin.id
            ).count(),
            0,
        )
        self.assertEqual(
            self.db.query(LeaveRequest).filter(LeaveRequest.user_id == admin.id).count(),
            0,
        )
        self.assertEqual(
            self.db.query(UserPermission).filter(
                UserPermission.user_id == admin.id
            ).count(),
            0,
        )
        self.assertEqual(
            self.db.query(Notification).filter(Notification.user_id == admin.id).count(),
            0,
        )
        self.assertEqual(
            self.db.query(Event).filter(Event.created_by == admin.id).count(),
            0,
        )
        self.assertEqual(
            self.db.query(RoomReservation).filter(
                RoomReservation.created_by == admin.id
            ).count(),
            0,
        )
        self.assertNotIn(admin.id, scoped_user_ids(self.db, super_admin))

    def test_purge_inactive_users_removes_existing_passive_accounts(self):
        inactive = self.add_user("Inactive", "inactive@example.com", "employee")
        inactive.is_active = False
        self.db.commit()

        deleted_count = purge_inactive_users(self.db)
        self.db.commit()

        self.assertEqual(deleted_count, 1)
        self.assertIsNone(self.db.query(User).filter(User.id == inactive.id).first())

    def test_purge_old_menus_removes_entries_older_than_three_days(self):
        today = turkey_today()
        old_date = today - timedelta(days=4)
        boundary_date = today - timedelta(days=3)
        old_menu = Menu(menu_date=old_date, content="Eski")
        boundary_menu = Menu(menu_date=boundary_date, content="Sınır")
        today_menu = Menu(menu_date=today, content="Bugün")
        self.db.add_all([old_menu, boundary_menu, today_menu])
        self.db.commit()

        deleted_count = purge_old_menus(self.db)
        remaining_dates = {
            menu.menu_date
            for menu in self.db.query(Menu).all()
        }

        self.assertEqual(deleted_count, 1)
        self.assertNotIn(old_date, remaining_dates)
        self.assertIn(boundary_date, remaining_dates)
        self.assertIn(today, remaining_dates)

    def test_excel_export_excludes_users_without_work_hours(self):
        super_admin = self.add_user("Super", "super-excel@example.com", "super_admin")
        scheduled = self.add_user("Scheduled", "scheduled@example.com", "employee")
        unscheduled = self.add_user("Unscheduled", "unscheduled@example.com", "employee")
        scheduled.work_start_time = time(9, 0)
        scheduled.work_end_time = time(18, 0)
        self.db.commit()

        response = export_attendance_excel(
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        async def read_body():
            return b"".join([chunk async for chunk in response.body_iterator])

        body = asyncio.run(read_body())
        workbook = load_workbook(BytesIO(body))
        names = {
            row[1]
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[1]
        }

        self.assertIn("Scheduled", names)
        self.assertNotIn("Unscheduled", names)

    def test_user_permission_updates_are_idempotent(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        employee = self.add_user("Employee", "employee@example.com", "employee")
        permission = Permission(
            code="menu.manage",
            description="Menu management",
        )
        self.db.add(permission)
        self.db.commit()

        payload = UserPermissionCreate(permission_code=permission.code)

        first_assign = assign_permission_to_user(
            employee.id,
            payload,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        second_assign = assign_permission_to_user(
            employee.id,
            payload,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        first_remove = remove_permission_from_user(
            employee.id,
            permission.code,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        second_remove = remove_permission_from_user(
            employee.id,
            permission.code,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )

        self.assertEqual(first_assign["permission"], permission.code)
        self.assertEqual(second_assign["permission"], permission.code)
        self.assertEqual(first_remove["permission"], permission.code)
        self.assertEqual(second_remove["permission"], permission.code)

    def test_user_permission_update_creates_missing_permission_definition(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        employee = self.add_user("Employee", "employee@example.com", "employee")
        payload = UserPermissionCreate(permission_code="leave.approve")

        response = assign_permission_to_user(
            employee.id,
            payload,
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )
        permission = self.db.query(Permission).filter(
            Permission.code == payload.permission_code
        ).first()

        self.assertEqual(response["permission"], payload.permission_code)
        self.assertIsNotNone(permission)

    def test_remove_missing_permission_definition_is_successful(self):
        super_admin = self.add_user("Super", "super@example.com", "super_admin")
        employee = self.add_user("Employee", "employee@example.com", "employee")

        response = remove_permission_from_user(
            employee.id,
            "attendance.view",
            {"sub": super_admin.email, "role": super_admin.role},
            self.db,
        )

        self.assertEqual(response["permission"], "attendance.view")

    def test_create_event_notifies_all_active_users(self):
        creator = self.add_user("Creator", "creator@example.com", "super_admin")
        employee = self.add_user("Employee", "employee@example.com", "employee")
        inactive = self.add_user("Inactive", "inactive-event@example.com", "employee")
        inactive.is_active = False
        self.db.commit()

        response = create_event(
            EventCreate(
                title="Toplantı",
                description="Genel toplantı",
                location="Salon",
                start_time=datetime.utcnow() + timedelta(days=1),
            ),
            {"sub": creator.email, "role": creator.role},
            self.db,
        )
        notified_user_ids = {
            notification.user_id
            for notification in self.db.query(Notification).filter(
                Notification.link == "/events",
            )
        }

        self.assertIsNotNone(response["event_id"])
        self.assertEqual(notified_user_ids, {creator.id, employee.id})

    def test_event_reminder_notifies_once_three_hours_before_start(self):
        creator = self.add_user("Creator", "creator-reminder@example.com", "super_admin")
        employee = self.add_user("Employee", "employee-reminder@example.com", "employee")
        due_event = Event(
            title="Yakın Etkinlik",
            start_time=turkey_now() + timedelta(hours=2, minutes=55),
            created_by=creator.id,
        )
        later_event = Event(
            title="Sonraki Etkinlik",
            start_time=turkey_now() + timedelta(hours=4),
            created_by=creator.id,
        )
        self.db.add_all([due_event, later_event])
        self.db.commit()

        sent_count = send_due_event_reminders(self.db)
        second_sent_count = send_due_event_reminders(self.db)
        reminder_notifications = self.db.query(Notification).filter(
            Notification.title == "Etkinlik hatırlatması",
            Notification.link == "/events",
        ).all()

        self.db.refresh(due_event)
        self.db.refresh(later_event)

        self.assertEqual(sent_count, 1)
        self.assertEqual(second_sent_count, 0)
        self.assertIsNotNone(due_event.reminder_sent_at)
        self.assertIsNone(later_event.reminder_sent_at)
        self.assertEqual(
            {notification.user_id for notification in reminder_notifications},
            {creator.id, employee.id},
        )


if __name__ == "__main__":
    unittest.main()
