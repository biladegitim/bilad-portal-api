from datetime import datetime, timedelta
from calendar import monthrange
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database.connection import get_db
from app.models.qr import QRToken
from app.models.attendance import AttendanceRecord
from app.models.leave import LeaveRequest
from app.models.user import User
from app.schemas.attendance import AttendanceScan
from app.core.dependencies import admin_required, get_current_user, qr_display_required
from app.core.rbac import get_db_user_from_token, normalize_role, scoped_users_query


router = APIRouter()


def each_date(start_date, end_date):
    current_date = start_date

    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)


def serialize_daily_report(user: User, records: list[AttendanceRecord]):
    daily_records = {}

    for record in records:
        record_date = record.record_time.date()
        daily_records.setdefault(record_date, []).append(record)

    user_report = []

    for record_date, day_records in daily_records.items():
        first_entry = None
        last_exit = None

        for record in day_records:
            if record.record_type == "check_in" and not first_entry:
                first_entry = record.record_time

            if record.record_type == "check_out":
                last_exit = record.record_time

        late = False
        early_exit = False

        if user.work_start_time and first_entry:
            expected_start = datetime.combine(record_date, user.work_start_time)
            late = first_entry > expected_start

        if user.work_end_time and last_exit:
            expected_end = datetime.combine(record_date, user.work_end_time)
            early_exit = last_exit < expected_end

        user_report.append({
            "date": str(record_date),
            "first_entry": first_entry,
            "last_exit": last_exit,
            "late": late,
            "early_exit": early_exit,
        })

    return user_report


@router.get("/qr/current")
def get_current_qr(
    current_user: dict = Depends(qr_display_required),
    db: Session = Depends(get_db),
):
    qr_token = QRToken()

    db.add(qr_token)
    db.commit()
    db.refresh(qr_token)

    return {
        "token": qr_token.token,
        "expires_at": qr_token.expires_at,
    }


@router.post("/attendance/scan")
def scan_attendance(
    data: AttendanceScan,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    qr_token = db.query(QRToken).filter(QRToken.token == data.token).first()

    if not qr_token:
        raise HTTPException(status_code=404, detail="QR kod bulunamadÄ±")

    if qr_token.is_used:
        raise HTTPException(status_code=400, detail="Bu QR kod daha Ã¶nce kullanÄ±lmÄ±ÅŸ")

    if qr_token.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="QR kodun sÃ¼resi dolmuÅŸ")

    user = get_db_user_from_token(db, current_user)

    if normalize_role(user.role) != "super_admin":
        if not user.device_id:
            user.device_id = data.device_id
        elif user.device_id != data.device_id:
            raise HTTPException(
                status_code=403,
                detail="Bu hesap farklÄ± bir cihaza tanÄ±mlÄ±",
            )

    today = datetime.utcnow().date()
    today_start = datetime.combine(today, datetime.min.time())

    last_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.user_id == user.id,
        AttendanceRecord.record_time >= today_start,
    ).order_by(
        AttendanceRecord.record_time.desc()
    ).first()

    record_type = "check_out" if last_record and last_record.record_type == "check_in" else "check_in"

    attendance = AttendanceRecord(
        user_id=user.id,
        record_type=record_type,
        record_time=datetime.utcnow(),
        source="dynamic_qr",
    )

    qr_token.is_used = True

    db.add(attendance)
    db.commit()
    db.refresh(attendance)

    return {
        "message": "GiriÅŸ-Ã§Ä±kÄ±ÅŸ kaydÄ± oluÅŸturuldu",
        "record_type": record_type,
        "record_time": attendance.record_time,
    }


@router.get("/attendance/today")
def get_today_attendance(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user = get_db_user_from_token(db, current_user)
    today = datetime.utcnow().date()

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.user_id == user.id,
        AttendanceRecord.record_time >= datetime.combine(today, datetime.min.time()),
        AttendanceRecord.record_time <= datetime.combine(today, datetime.max.time()),
    ).order_by(
        AttendanceRecord.record_time.asc()
    ).all()

    return {
        "date": str(today),
        "user": {
            "id": user.id,
            "full_name": user.full_name,
        },
        "records": [
            {
                "id": record.id,
                "record_type": record.record_type,
                "record_time": record.record_time,
                "source": record.source,
            }
            for record in records
        ],
    }


@router.get("/attendance/weekly-report")
def get_weekly_report(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    current_db_user = get_db_user_from_token(db, current_user)
    users = scoped_users_query(db, current_db_user).order_by(User.id.asc()).all()

    today = datetime.utcnow().date()
    week_ago = today - timedelta(days=7)
    report_data = []

    for user in users:
        attendance_records = db.query(AttendanceRecord).filter(
            AttendanceRecord.user_id == user.id,
            AttendanceRecord.record_time >= datetime.combine(week_ago, datetime.min.time()),
        ).order_by(
            AttendanceRecord.record_time.asc()
        ).all()

        report_data.append({
            "user_id": user.id,
            "full_name": user.full_name,
            "position": user.position,
            "report": serialize_daily_report(user, attendance_records),
        })

    return {"weekly_report": report_data}


@router.get("/attendance/dashboard")
def get_attendance_dashboard(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    current_db_user = get_db_user_from_token(db, current_user)
    users = scoped_users_query(db, current_db_user).order_by(User.id.asc()).all()
    user_ids = [user.id for user in users]

    today = datetime.utcnow().date()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())

    today_records = db.query(AttendanceRecord).filter(
        AttendanceRecord.user_id.in_(user_ids or [-1]),
        AttendanceRecord.record_time >= today_start,
        AttendanceRecord.record_time <= today_end,
    ).order_by(
        AttendanceRecord.record_time.asc()
    ).all()

    summary = []

    for user in users:
        user_records = [
            record for record in today_records
            if record.user_id == user.id
        ]

        check_ins = [
            record for record in user_records
            if record.record_type == "check_in"
        ]

        check_outs = [
            record for record in user_records
            if record.record_type == "check_out"
        ]

        first_entry = check_ins[0].record_time if check_ins else None
        last_exit = check_outs[-1].record_time if check_outs else None
        is_inside = bool(user_records and user_records[-1].record_type == "check_in")
        late = False

        if first_entry and user.work_start_time:
            expected_start = datetime.combine(today, user.work_start_time)
            late = first_entry > expected_start

        summary.append({
            "user_id": user.id,
            "full_name": user.full_name,
            "position": user.position,
            "first_entry": first_entry,
            "last_exit": last_exit,
            "is_inside": is_inside,
            "late": late,
        })

    inside_users = [item for item in summary if item["is_inside"]]
    arrived_users = [item for item in summary if item["first_entry"]]
    late_users = [item for item in summary if item["late"]]
    not_checked_out = [
        item for item in summary
        if item["first_entry"] and not item["last_exit"]
    ]

    return {
        "date": str(today),
        "counts": {
            "arrived": len(arrived_users),
            "inside": len(inside_users),
            "late": len(late_users),
            "not_checked_out": len(not_checked_out),
        },
        "inside_users": inside_users,
        "late_users": late_users,
        "not_checked_out": not_checked_out,
        "summary": summary,
    }


@router.get("/attendance/export-excel")
def export_attendance_excel(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    current_db_user = get_db_user_from_token(db, current_user)
    users = scoped_users_query(db, current_db_user).order_by(User.full_name.asc()).all()
    user_ids = [user.id for user in users]

    now_utc = datetime.utcnow()
    turkey_offset = timedelta(hours=3)
    now_local = now_utc + turkey_offset
    tolerance = timedelta(minutes=10)

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.user_id.in_(user_ids or [-1]),
        AttendanceRecord.record_time <= now_utc,
    ).order_by(
        AttendanceRecord.record_time.asc()
    ).all()

    records_by_user_day = {}

    for record in records:
        local_time = record.record_time + turkey_offset
        key = (record.user_id, local_time.date())
        records_by_user_day.setdefault(key, []).append((record, local_time))

    today = now_local.date()
    today_end_local = datetime.combine(today, datetime.max.time())
    approved_leaves = db.query(LeaveRequest).filter(
        LeaveRequest.user_id.in_(user_ids or [-1]),
        LeaveRequest.status == "approved",
        LeaveRequest.start_time <= today_end_local,
    ).all()
    leaves_by_user_day = {}

    for leave in approved_leaves:
        leave_start = leave.start_time.date()
        leave_end = min(leave.end_time.date(), today)

        if leave_end < leave_start:
            continue

        for leave_date in each_date(leave_start, leave_end):
            leaves_by_user_day.setdefault((leave.user_id, leave_date), []).append(leave)

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="1F2937")
    warning_fill = PatternFill("solid", fgColor="FEE2E2")
    warning_font = Font(bold=True, color="B91C1C")
    leave_fill = PatternFill("solid", fgColor="DBEAFE")
    leave_font = Font(bold=True, color="1D4ED8")
    center_alignment = Alignment(horizontal="center", vertical="center")
    month_names = {
        1: "Ocak",
        2: "Subat",
        3: "Mart",
        4: "Nisan",
        5: "Mayis",
        6: "Haziran",
        7: "Temmuz",
        8: "Agustos",
        9: "Eylul",
        10: "Ekim",
        11: "Kasim",
        12: "Aralik",
    }

    report_months = sorted({
        (record_date.year, record_date.month)
        for _, record_date in [
            *records_by_user_day.keys(),
            *leaves_by_user_day.keys(),
        ]
    })
    current_month = (today.year, today.month)

    if current_month not in report_months:
        report_months.append(current_month)

    wb = Workbook()
    wb.remove(wb.active)

    for year, month in report_months:
        ws = wb.create_sheet(title=f"{month_names[month]} {year}")
        ws.append(["Tarih", "Ad Soyad", "Giriş Saati", "Çıkış Saati", "Durum"])

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        row_index = 2

        month_start = datetime(year, month, 1).date()
        month_end = datetime(year, month, monthrange(year, month)[1]).date()

        if (year, month) == current_month:
            month_end = today

        for user in users:
            user_days = list(each_date(month_start, month_end))

            for record_date in user_days:
                day_records = records_by_user_day.get((user.id, record_date), [])
                approved_day_leaves = leaves_by_user_day.get((user.id, record_date), [])
                has_approved_leave = bool(approved_day_leaves)

                if not day_records and record_date.weekday() == 6 and not has_approved_leave:
                    continue

                if (
                    not day_records
                    and not has_approved_leave
                    and record_date == today
                ):
                    if not user.work_end_time:
                        continue

                    expected_end_today = datetime.combine(record_date, user.work_end_time)

                    if now_local <= expected_end_today + tolerance:
                        continue

                check_ins = [
                    local_time
                    for record, local_time in day_records
                    if record.record_type == "check_in"
                ]
                check_outs = [
                    local_time
                    for record, local_time in day_records
                    if record.record_type == "check_out"
                ]

                first_entry = check_ins[0] if check_ins else None
                last_exit = check_outs[-1] if check_outs else None
                status = ""

                if has_approved_leave:
                    status = "İzinli"
                elif not day_records:
                    status = "Gelmedi"

                ws.append([
                    record_date.strftime("%d.%m.%Y"),
                    user.full_name,
                    first_entry.strftime("%H:%M") if first_entry else "-",
                    last_exit.strftime("%H:%M") if last_exit else "-",
                    status,
                ])

                entry_cell = ws.cell(row=row_index, column=3)
                exit_cell = ws.cell(row=row_index, column=4)
                status_cell = ws.cell(row=row_index, column=5)

                if user.work_start_time and first_entry:
                    expected_start = datetime.combine(record_date, user.work_start_time)
                    if first_entry > expected_start + tolerance:
                        entry_cell.fill = warning_fill
                        entry_cell.font = warning_font

                if user.work_end_time and last_exit:
                    expected_end = datetime.combine(record_date, user.work_end_time)
                    if last_exit < expected_end - tolerance:
                        exit_cell.fill = warning_fill
                        exit_cell.font = warning_font

                if status == "Gelmedi":
                    status_cell.fill = warning_fill
                    status_cell.font = warning_font
                elif status == "İzinli":
                    status_cell.fill = leave_fill
                    status_cell.font = leave_font

                for cell in ws[row_index]:
                    cell.alignment = center_alignment

                row_index += 1

        widths = [14, 28, 16, 16, 16]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        if (year, month) == current_month:
            wb.active = wb.worksheets.index(ws)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"giris-cikis-raporu-{today}.xlsx"

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
